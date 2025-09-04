# encoding: utf-8
from __future__ import annotations

import ckan.plugins as p
from ckanext.datastore.interfaces import IDatastoreDump
from typing import Any
from contextlib import contextmanager
from io import BytesIO


@contextmanager
def xlsx_writer(fields: list[dict[str, Any]], bom: bool = False):
    """Context manager for writing XLSX data to file

    :param fields: list of datastore fields
    :param bom: ignored for XLSX format
    """
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl is required for XLSX export. "
                          "Install it with: pip install openpyxl")

    output = BytesIO()
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Data"

    # Write headers
    for idx, field in enumerate(fields, 1):
        col_letter = get_column_letter(idx)
        worksheet[f'{col_letter}1'] = field['id']

    yield XLSXWriter(output, workbook, worksheet, len(fields))


class XLSXWriter(object):
    def __init__(self, output: BytesIO, workbook: Any,
                 worksheet: Any, num_fields: int):
        self.output = output
        self.workbook = workbook
        self.worksheet = worksheet
        self.num_fields = num_fields
        self.current_row = 2  # Start after header row

    def write_records(self, records: list[dict[str, Any]]) -> bytes:
        """Write records to the XLSX worksheet"""
        try:
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("openpyxl is required for XLSX export")

        for record in records:
            for idx in range(self.num_fields):
                col_letter = get_column_letter(idx + 1)
                # Get the field name from the first record if available
                field_names = list(record.keys()) if record else []
                if idx < len(field_names):
                    field_name = field_names[idx]
                    value = record.get(field_name, '')
                    self.worksheet[f'{col_letter}{self.current_row}'] = value
            self.current_row += 1

        return b''  # No incremental output for XLSX

    def end_file(self) -> bytes:
        """Finalize the XLSX file and return the complete file"""
        self.workbook.save(self.output)
        self.output.seek(0)
        return self.output.read()


class ExampleXLSXPlugin(p.SingletonPlugin):
    """Example plugin that adds XLSX export capability to datastore"""

    p.implements(IDatastoreDump)

    def register_dump_formats(self) -> dict[str, dict[str, Any]]:
        """Register XLSX format for datastore exports"""
        return {
            'xlsx': {
                'writer_factory': xlsx_writer,
                'records_format': 'objects',
                'content_type': (b'application/vnd.openxmlformats-'
                                 b'officedocument.spreadsheetml.sheet'),
                'file_extension': 'xlsx'
            }
        }
